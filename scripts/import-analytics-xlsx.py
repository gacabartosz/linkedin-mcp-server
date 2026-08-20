#!/usr/bin/env python3
"""
Importer eksportu LinkedIn "AggregateAnalytics" (.xlsx) → analytics.db (creator_analytics).

Zastępuje kruchy web-scrape Highcharts (LinkedIn przebudował stronę — window.Highcharts zniknęło).
Flow: user klika "Export" na LinkedIn Analytics → plik ląduje w ~/Downloads → ten skrypt go parsuje
i wypełnia 365-dniowe serie, które czyta dashboard (dashboard.mjs:1849-1853).

Serie (kontrakt dashboardu): followers/cumulative, followers/daily,
                             impressions/cumulative, impressions/daily, engagements/daily

Użycie:
  python3 scripts/import-analytics-xlsx.py [ścieżka.xlsx]
  (bez argumentu: najnowszy ~/Downloads/AggregateAnalytics_*.xlsx)
"""
import sys, os, glob, sqlite3, datetime, re
import openpyxl

HOME = os.path.expanduser("~")
DB = os.path.join(HOME, ".linkedin-mcp", "analytics.db")


def newest_export():
    files = glob.glob(os.path.join(HOME, "Downloads", "AggregateAnalytics_*.xlsx"))
    if not files:
        sys.exit("Brak plików ~/Downloads/AggregateAnalytics_*.xlsx — zrób najpierw Export na LinkedIn.")
    return max(files, key=os.path.getmtime)


def iso(d):
    if isinstance(d, (datetime.datetime, datetime.date)):
        return d.strftime("%Y-%m-%d")
    s = str(d).strip()
    if "." in s:
        day, mon, yr = (s.split(".") + ["", "", ""])[:3]
        return f"{int(yr):04d}-{int(mon):02d}-{int(day):02d}"
    return s


def num(v):
    if v is None:
        return None
    s = str(v).replace(" ", "").replace(" ", "").replace(",", "")
    try:
        return int(float(s))
    except ValueError:
        return None


def header_row(rows):
    for i, r in enumerate(rows):
        if r and r[0] is not None and str(r[0]).strip().lower() == "data":
            return i
    return None


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else newest_export()
    print(f"Import: {path}")
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    # OBSERWUJĄCY → total followers + dzienni "Nowi obserwujący"
    rows = list(wb["OBSERWUJĄCY"].iter_rows(values_only=True))
    total_followers = num(rows[0][1])
    h = header_row(rows)
    foll_daily = [(iso(r[0]), num(r[1])) for r in rows[h + 1:] if r and r[0] is not None and num(r[1]) is not None]

    # REAKCJE → dzienne Wyświetlenia (impressions) + Reakcje (engagements)
    rrows = list(wb["REAKCJE"].iter_rows(values_only=True))
    rh = header_row(rrows)
    impr_daily, eng_daily = [], []
    for r in rrows[rh + 1:]:
        if not r or r[0] is None:
            continue
        d = iso(r[0])
        iv = num(r[1])
        rv = num(r[2]) if len(r) > 2 else None
        if iv is not None:
            impr_daily.append((d, iv))
        if rv is not None:
            eng_daily.append((d, rv))

    # followers/cumulative = ABSOLUTNA liczba: baseline(start) = total − suma przyrostów
    sum_new = sum(v for _, v in foll_daily)
    baseline = (total_followers - sum_new) if total_followers is not None else 0
    foll_cum, run = [], baseline
    for d, v in foll_daily:
        run += v
        foll_cum.append((d, run))
    # impressions/cumulative = suma narastająca od 0
    impr_cum, run = [], 0
    for d, v in impr_daily:
        run += v
        impr_cum.append((d, run))

    series = [
        ("followers", "cumulative", foll_cum),
        ("followers", "daily", foll_daily),
        ("impressions", "cumulative", impr_cum),
        ("impressions", "daily", impr_daily),
        ("engagements", "daily", eng_daily),
    ]

    con = sqlite3.connect(DB)
    con.execute("""CREATE TABLE IF NOT EXISTS creator_analytics (
        date TEXT NOT NULL, metric TEXT NOT NULL, chart_type TEXT NOT NULL,
        value INTEGER, scraped_at TEXT DEFAULT (datetime('now')),
        PRIMARY KEY (date, metric, chart_type))""")
    now = datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    n = 0
    for metric, ctype, data in series:
        for d, v in data:
            con.execute(
                "INSERT OR REPLACE INTO creator_analytics (date,metric,chart_type,value,scraped_at) VALUES (?,?,?,?,?)",
                (d, metric, ctype, v, now))
            n += 1
    # === Top posty (NAJPOPULARNIEJSZE PUBLIKACJE) → xlsx_top_posts ===
    # Prawdziwe, kompletne top-posty z oficjalnego eksportu (social_metadata jest niepełne i zaniża).
    def act_id(u):
        m = re.search(r"(\d{15,})", str(u or ""))
        return m.group(1) if m else None
    prows = list(wb["NAJPOPULARNIEJSZE PUBLIKACJE"].iter_rows(values_only=True))
    ph = next((i for i, r in enumerate(prows) if r and str(r[0]).strip().startswith("Adres URL")), 0)
    con.execute("DROP TABLE IF EXISTS xlsx_top_posts")
    con.execute("""CREATE TABLE xlsx_top_posts (
        kind TEXT, rank INTEGER, post_url TEXT, activity_id TEXT, post_date TEXT, value INTEGER, imported_at TEXT)""")
    rr = ri = 0
    for r in prows[ph + 1:]:
        if not r:
            continue
        if r[0] and str(r[0]).startswith("http") and num(r[2]) is not None:
            rr += 1
            con.execute("INSERT INTO xlsx_top_posts VALUES ('reactions',?,?,?,?,?,?)",
                        (rr, str(r[0]), act_id(r[0]), iso(r[1]), num(r[2]), now))
        if len(r) > 4 and r[4] and str(r[4]).startswith("http") and num(r[6]) is not None:
            ri += 1
            con.execute("INSERT INTO xlsx_top_posts VALUES ('impressions',?,?,?,?,?,?)",
                        (ri, str(r[4]), act_id(r[4]), iso(r[5]), num(r[6]), now))

    # === Weekly: realne reakcje + przyrost followers z dziennych serii (tygodnie niedziela-start, jak dashboard) ===
    from datetime import date as _date, timedelta
    from collections import defaultdict

    def wk(iso_d):
        y, m, dd = map(int, iso_d.split("-"))
        dt = _date(y, m, dd)
        return (dt - timedelta(days=(dt.weekday() + 1) % 7)).isoformat()

    wkr, wki, wkf = defaultdict(int), defaultdict(int), defaultdict(int)
    for d, v in eng_daily:
        wkr[wk(d)] += v
    for d, v in impr_daily:
        wki[wk(d)] += v
    for d, v in foll_daily:
        wkf[wk(d)] += v
    wkp = defaultdict(int)  # posty/tydzień ze scheduler.db
    try:
        scon = sqlite3.connect(os.path.join(HOME, ".linkedin-mcp", "scheduler.db"))
        for (pa,) in scon.execute("SELECT publish_at FROM scheduled_posts WHERE status='published' AND publish_at IS NOT NULL"):
            try:
                wkp[wk(str(pa)[:10])] += 1
            except Exception:
                pass
        scon.close()
    except Exception as e:
        print("weekly posts warn:", e)
    con.execute("""CREATE TABLE IF NOT EXISTS weekly_report (
        week_start TEXT PRIMARY KEY, top_post_urn TEXT, top_post_reactions INTEGER,
        total_reactions INTEGER, follower_delta INTEGER, posts_count INTEGER,
        avg_engagement_rate REAL, content_mix TEXT, report_text TEXT)""")
    wk_n = 0
    for ws_ in sorted(set(wkr) | set(wkf) | set(wkp)):
        tr, fd, pc = wkr.get(ws_, 0), wkf.get(ws_, 0), wkp.get(ws_, 0)
        eng = round(tr / wki[ws_] * 100, 1) if wki.get(ws_) else 0
        con.execute("""INSERT INTO weekly_report (week_start, total_reactions, follower_delta, posts_count, avg_engagement_rate)
            VALUES (?,?,?,?,?)
            ON CONFLICT(week_start) DO UPDATE SET total_reactions=excluded.total_reactions,
              follower_delta=excluded.follower_delta, posts_count=excluded.posts_count,
              avg_engagement_rate=excluded.avg_engagement_rate""", (ws_, tr, fd, pc, eng))
        wk_n += 1

    con.commit()
    rng = con.execute("SELECT min(date),max(date),count(*) FROM creator_analytics").fetchone()
    tp_cnt = con.execute("SELECT count(*) FROM xlsx_top_posts WHERE kind='reactions'").fetchone()[0]
    tp_top = con.execute("SELECT value FROM xlsx_top_posts WHERE kind='reactions' ORDER BY rank LIMIT 1").fetchone()
    con.close()
    print(f"Top posty (reakcje): {tp_cnt}, najlepszy = {tp_top[0] if tp_top else '—'} reakcji.")

    # stamp last_import w scraper-schedule.json (źródło dla panelu w dashboardzie)
    try:
        import json
        sp = os.path.join(HOME, ".linkedin-mcp", "scraper-schedule.json")
        cfg = json.load(open(sp, encoding="utf-8")) if os.path.exists(sp) else {}
        cfg.setdefault("charts_365d", {})
        cfg["charts_365d"]["last_import"] = now
        cfg["charts_365d"]["last_import_file"] = os.path.basename(path)
        json.dump(cfg, open(sp, "w", encoding="utf-8"), ensure_ascii=False, indent=2)
    except Exception as e:
        print("stamp schedule warn:", e)

    print(f"Zapisano {n} punktów. creator_analytics: {rng[2]} wierszy, {rng[0]} → {rng[1]}.")
    print(f"Total followers={total_followers}, baseline(start)={baseline}, suma przyrostów={sum_new}.")
    for metric, ctype, data in series:
        if data:
            print(f"  {metric}/{ctype}: {len(data)} pkt, ostatni {data[-1]}")


if __name__ == "__main__":
    main()
